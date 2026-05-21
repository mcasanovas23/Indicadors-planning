"""
Dashboard Indicadors Planning - iVascular
Reads Excel → computes all indicators → generates index.html → git push
"""
import json
import re
import shutil
import subprocess
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

# ── CONFIG ──────────────────────────────────────────────────────────────────
EXCEL_PATH = Path(r"C:\Users\mcasanovas\OneDrive - IVASCULAR, S.L.U\Indicadores Stents Marc.xlsm")
TEMP_EXCEL = Path(r"C:\Users\mcasanovas\stents_tmp.xlsx")
REPO_PATH  = Path(r"C:\Users\mcasanovas\OneDrive - IVASCULAR, S.L.U\Código herramientas")
HTML_OUT   = REPO_PATH / "index.html"
EXCEL_OUT  = REPO_PATH / "datos.xlsx"

FAMILY_ORDER = ["IVOLUTIO", "ANGIOLIT", "NAVISCOR", "RESTORER", "ICOVER", "INTERCEP"]
FAMILY_NAMES = {
    "ICOVER":   "ICOVER",
    "ANGIOLIT": "ANGIOLITE",
    "NAVISCOR": "NAVISCORE",
    "IVOLUTIO": "IVOLUTION",
    "RESTORER": "RESTORER",
    "INTERCEP": "INTERCEPT",
}
PA_TO_DASH = {
    "ANGIOLIT": "ANGIOLIT", "ANGIOBTK": "ANGIOLIT",
    "ICOVER":   "ICOVER",   "TUVA":     "ICOVER",
    "IVOLPRO":  "IVOLUTIO", "IVOSPRO":  "IVOLUTIO", "IVOXLPRO": "IVOLUTIO",
    "NAVISCOR": "NAVISCOR",
    "RESTORER": "RESTORER",
    "INTERCEP": "INTERCEP",
}

FACTORY_ORDER = ["NITI", "DES", "SCCD", "BCSR", "PERI", "IVPG_E", "IVPG_B", "CESTA"]
FACTORY_LABELS = {
    "NITI":   "NITI (Ivolutio)",
    "DES":    "DES (Angiolit DES)",
    "SCCD":   "SCCD (Bare Angiolit)",
    "BCSR":   "BCSR (Naviscor)",
    "PERI":   "PERI (Restorer)",
    "IVPG_E": "IVPG_E (Encapsulado)",
    "IVPG_B": "IVPG_B (Bare Icover)",
    "CESTA":  "CESTA (Intercep)",
}
def art_to_factory(art):
    a = str(art)
    if a.startswith("SE221"): return "IVPG_B"   # before SE22 generic
    if a.startswith("SE220"): return "PERI"
    if a.startswith("SE21"):  return "SCCD"
    if a.startswith("SE32"):  return "DES"
    if a.startswith("SE23"):  return "IVPG_E"
    if a.startswith("SE24"):  return "NITI"
    if a.startswith("SE26"):  return "BCSR"
    if a.startswith("SE28"):  return "CESTA"
    return None

# SE article → dashboard family (by prefix)
def se_to_family(article):
    a = str(article)
    if a.startswith("SE21"): return "ANGIOLIT"
    if a.startswith("SE220"): return "RESTORER"
    if a.startswith("SE221"): return "ICOVER"
    if a.startswith("SE23"): return "ICOVER"
    if a.startswith("SE24"): return "IVOLUTIO"
    if a.startswith("SE26"): return "NAVISCOR"
    if a.startswith("SE28"): return "INTERCEP"
    if a.startswith("SE32"): return "ANGIOLIT"
    return None

# Layers per family: (title, article_filter_fn, offset, is_bare)
LAYERS = {
    "ICOVER": [
        ("STENT ENCAPSULADO (SE23)",       lambda a: a.startswith("SE23"),  1, False),
        ("STENT BARE (SE22 ICOVER)",        lambda a: a.startswith("SE221"), 1, True),
    ],
    "ANGIOLIT": [
        ("STENT RECUBIERTO DES (SE32)",     lambda a: a.startswith("SE32"),  1, False),
        ("STENT BARE DES (SE21)",           lambda a: a.startswith("SE21"),  3, True),
    ],
    "NAVISCOR": [
        ("STENT NAVISCOR",                  lambda a: a.startswith("SE26"),  1, False),
    ],
    "IVOLUTIO": [
        ("STENT IVOLUTIO",                  lambda a: a.startswith("SE24"),  1, False),
    ],
    "RESTORER": [
        ("STENT RESTORER",                  lambda a: a.startswith("SE220"), 1, False),
    ],
    "INTERCEP": [
        ("STENT INTERCEP",                  lambda a: a.startswith("SE28"),  1, False),
    ],
}

# ── MEDIDA PARSING ──────────────────────────────────────────────────────────
def parse_medida(nom_art):
    """Extract abbreviated measurement from Fases_SE nom_art description."""
    if not nom_art:
        return ""
    s = str(nom_art)
    # SE220 RESTORER: "Stent Periférico Small 18" → "S x 18"
    m = re.search(r'\b(Small|Medium|Large)\s+(\d+)\b', s, re.IGNORECASE)
    if m:
        return f"{m.group(1)[0].upper()} x {m.group(2)}"
    # SE221/SE23: "5 a 8 x 17 mm" → "5-8 x 17"
    m = re.search(r'(\d+)\s+a\s+(\d+)\s+x\s+(\d+)', s, re.IGNORECASE)
    if m:
        return f"{m.group(1)}-{m.group(2)} x {m.group(3)}"
    # SE21/SE32: "2,00 a 2,50mm* 9 mm" or "2 a 2,50mm* 9 mm" → "2,00-2,50 x 9"
    m = re.search(r'(\d+(?:[,\.]\d+)?)\s*a\s*(\d+(?:[,\.]\d+)?)\s*mm\*?\s*(\d+)', s, re.IGNORECASE)
    if m:
        return f"{m.group(1)}-{m.group(2)} x {m.group(3)}"
    # SE24 IVOLUTIO: "5x40 mm" → "5 x 40"
    m = re.search(r'(\d+)x(\d+)\s*mm', s, re.IGNORECASE)
    if m:
        return f"{m.group(1)} x {m.group(2)}"
    # SE26 NAVISCOR: "D1,5*L6mm" → "1,5 x 6"
    m = re.search(r'D(\d+[,\.]?\d*)\*L(\d+)', s, re.IGNORECASE)
    if m:
        return f"{m.group(1)} x {m.group(2)}"
    # SE28 INTERCEP: "3*15" → "3 x 15"
    m = re.search(r'(\d+)\*(\d+)', s)
    if m:
        return f"{m.group(1)} x {m.group(2)}"
    return ""

def get_fases_medida(art, fases_medida):
    """Get abbreviated medida for an article from Fases_SE lookup."""
    nom = fases_medida.get(art)
    if nom:
        return parse_medida(nom)
    # Fallback: try without _T suffix
    if art.endswith("_T"):
        nom = fases_medida.get(art[:-2])
        if nom:
            return parse_medida(nom)
    return ""

# ── HELPERS ─────────────────────────────────────────────────────────────────
def date_to_week(d):
    """Convert a date to the company's week number (Fri-Sun → next ISO week)."""
    if d is None:
        return None
    if isinstance(d, str):
        try:
            from datetime import datetime
            d = datetime.strptime(d[:10], "%Y-%m-%d").date()
        except Exception:
            return None
    if hasattr(d, "date"):
        d = d.date()
    if d.weekday() >= 4:
        d = d + timedelta(days=7 - d.weekday())
    return d.isocalendar()[1]

def current_week():
    return date_to_week(date.today())

def load_excel():
    """Copy Excel (handles open-file lock) then load."""
    try:
        shutil.copy2(EXCEL_PATH, TEMP_EXCEL)
    except PermissionError:
        pass  # use existing temp copy
    wb = openpyxl.load_workbook(TEMP_EXCEL, read_only=True, data_only=True)
    def sheet_to_rows(name):
        ws = wb[name]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        return [dict(zip(headers, row)) for row in ws.iter_rows(min_row=2, values_only=True)]
    pa    = sheet_to_rows("Planning_PA")
    se    = sheet_to_rows("Planning_SE")
    inv   = sheet_to_rows("Inventari")
    fases = sheet_to_rows("Fases_SE")
    # Build medida lookup from Fases_SE cols G (cod_art) and H (nom_art)
    fases_medida = {}
    ws_f = wb["Fases_SE"]
    for row in ws_f.iter_rows(min_row=2, values_only=True):
        if len(row) >= 8 and row[6] and row[7] and row[6] != "cod_art":
            fases_medida[str(row[6])] = str(row[7])
    wb.close()
    return pa, se, inv, fases, fases_medida

# ── INVENTORY ───────────────────────────────────────────────────────────────
def build_inventory(inv_rows):
    """Return {article: available_units} using new cod_sta logic."""
    stock = defaultdict(int)
    for r in inv_rows:
        art = r.get("articulo")
        if not art:
            continue
        sta = r.get("cod_sta", 0)
        if sta == 0:
            stock[art] += (r.get("qnt_libre") or 0)
        elif sta == 40:
            stock[art] += (r.get("qnt_fis") or 0)
    return dict(stock)

# ── SE PLANNING ──────────────────────────────────────────────────────────────
def build_se_planning(se_rows):
    """Return {article: [{'semana':int, 'peces_lot':int, 'remaining':int, ...}]}"""
    plan = defaultdict(list)
    for r in se_rows:
        art = r.get("article")
        sem = r.get("setmana")
        lot = r.get("peces_lot") or 0
        if art and sem and lot > 0:
            plan[art].append({
                    "semana":    sem,
                    "peces_lot": lot,
                    "remaining": lot,
                    "lot":       r.get("lot"),
                    "fase":      r.get("fase_actual", ""),
                })
    for art in plan:
        plan[art].sort(key=lambda x: x["semana"])
    return dict(plan)

# ── PA ASSIGNMENT ────────────────────────────────────────────────────────────
STATUS_ASIGNADO   = "Asignado"
STATUS_EN_STOCK   = "En stock"
STATUS_EN_FAB     = "En fabricación"
STATUS_FABRICAR   = "Fabricar"

def run_pa_assignment(pa_rows, inventory, se_planning):
    """
    Enrich each PA row with 'status' and 'uds_fab'.
    Uses a running stock pool per SE article.
    """
    running_stock = defaultdict(int, inventory)
    # Deep-copy planning remainders
    import copy
    running_plan = {art: copy.deepcopy(lots) for art, lots in se_planning.items()}

    # Separate Asignado from those to process
    asignado, to_process = [], []
    for r in pa_rows:
        if r.get("disponibilitat_mat2") == STATUS_ASIGNADO:
            r["status"] = STATUS_ASIGNADO
            r["uds_fab"] = 0
            asignado.append(r)
        else:
            to_process.append(r)

    # Sort by semana, then original order (stable sort)
    to_process.sort(key=lambda r: (r.get("setmana") or 9999))

    for r in to_process:
        se_art = r.get("material2")
        qty    = r.get("peces_lot") or 0
        pa_sem = r.get("setmana") or 9999

        if not se_art or qty == 0:
            r["status"] = STATUS_ASIGNADO
            r["uds_fab"] = 0
            continue

        # Step 1: assign from stock
        from_stock = min(running_stock.get(se_art, 0), qty)
        running_stock[se_art] -= from_stock
        remaining = qty - from_stock

        if remaining == 0:
            r["status"] = STATUS_EN_STOCK
            r["uds_fab"] = 0
            continue

        # Step 2: assign from SE planning (semana < pa_sem)
        from_plan = 0
        for lot in running_plan.get(se_art, []):
            if lot["semana"] >= pa_sem:
                break
            take = min(lot["remaining"], remaining)
            lot["remaining"] -= take
            from_plan += take
            remaining  -= take
            if remaining == 0:
                break

        if remaining == 0:
            r["status"] = STATUS_EN_FAB
            r["uds_fab"] = 0
        else:
            r["status"]  = STATUS_FABRICAR
            r["uds_fab"] = remaining

    return asignado + to_process

# ── FAMILY COVERAGE MATRIX ───────────────────────────────────────────────────
def compute_family_matrix(pa_rows, se_rows, inventory, weeks, fases_articles=None, fases_medida=None):
    """
    Returns {family: {layer_title: [row_dict, ...]}}
    Each row_dict: {ref, medida, stock, en_curso, dem_6w, cob: [6 values], downstream}
    """
    w0 = weeks[0]

    # Build SE info: {article: {medida, familia}}
    se_info = {}
    for r in se_rows:
        art = r.get("article")
        if art and art not in se_info:
            se_info[art] = {"medida": r.get("medidas", ""), "familia": r.get("familia", "")}

    # SE planning by article: [(semana, peces_lot)]
    se_lots = defaultdict(list)
    for r in se_rows:
        art = r.get("article")
        sem = r.get("setmana")
        lot = r.get("peces_lot") or 0
        if art and sem and lot > 0:
            se_lots[art].append((sem, lot))

    # PA demand (non-Asignado) per SE article per semana
    pa_demand = defaultdict(lambda: defaultdict(int))
    for r in pa_rows:
        if r.get("disponibilitat_mat2") != STATUS_ASIGNADO:
            se_art = r.get("material2")
            sem    = r.get("setmana")
            qty    = r.get("peces_lot") or 0
            if se_art and sem:
                pa_demand[se_art][sem] += qty

    # Downstream mapping: SE21→SE32, SE22→SE23 via Planning_SE SRSI/encapsulado
    downstream = defaultdict(list)
    for r in se_rows:
        fam = r.get("familia", "")
        if fam in ("SRSI",):  # SE32 uses SE21 as material2
            art  = r.get("article")
            mat2 = r.get("material2")
            if art and mat2 and mat2 not in downstream[mat2]:
                downstream[mat2].append(art)

    def cob_array(art, stock_val, offset):
        result = []
        for w in weeks:
            plan = sum(qty for sem, qty in se_lots[art] if sem <= w - offset)
            dem  = sum(qty for sem, qty in pa_demand[art].items() if sem <= w)
            result.append(stock_val + plan - dem)
        return result

    # Collect all SE articles per family/layer
    # Source: Fases_SE col G (same as medida source) — only articles with known medida, no _T
    all_known = set(
        art for art in fases_medida.keys()
        if art and not art.endswith("_T")
    )
    fases_medida = fases_medida or {}
    result = {}
    for fam in FAMILY_ORDER:
        result[fam] = {}
        for title, filt, offset, is_bare in LAYERS[fam]:
            articles = sorted(set(art for art in all_known if filt(art)))
            rows = []
            for art in articles:
                stk    = inventory.get(art, 0)
                en_cur = sum(qty for sem, qty in se_lots[art] if w0 <= sem <= weeks[-1])
                dem6   = sum(qty for sem, qty in pa_demand[art].items() if sem in weeks)
                ds     = [a for a in downstream.get(art, []) if a in {r.get("article") for r in se_rows}]
                rows.append({
                    "ref":        art,
                    "medida":     get_fases_medida(art, fases_medida),
                    "stock":      stk,
                    "en_curso":   en_cur,
                    "dem_6w":     dem6,
                    "cob":        cob_array(art, stk, offset),
                    "downstream": ds,
                    "is_bare":    is_bare,
                })
            result[fam][title] = rows
    return result

# ── DASHBOARD DATA ────────────────────────────────────────────────────────────
def compute_dashboard(pa_enriched):
    """
    Returns {family: {pct: float, rows: [{se_ref, semana, qty}]}}
    rows = fabricar units grouped by (se_ref, semana), sorted by semana
    """
    # Group PA rows by dashboard family
    by_family = defaultdict(list)
    for r in pa_enriched:
        fam = PA_TO_DASH.get(r.get("familia", ""))
        if fam:
            by_family[fam].append(r)

    result = {}
    for fam in FAMILY_ORDER:
        rows_pa = by_family[fam]
        total   = len(rows_pa)
        covered = sum(1 for r in rows_pa if r["status"] in (STATUS_ASIGNADO, STATUS_EN_STOCK, STATUS_EN_FAB))
        pct     = round(covered / total * 100, 1) if total else 0

        # Fabricar rows grouped by (material2, semana)
        fab_agg = defaultdict(int)
        for r in rows_pa:
            if r["status"] == STATUS_FABRICAR:
                se_art = r.get("material2", "")
                sem    = r.get("setmana") or 0
                fab_agg[(se_art, sem)] += r.get("uds_fab", 0)

        # SE21/bare layer: from Planning_SE SRSI lots not yet assigned
        # (material2 of SRSI lots with disponibilitat_mat2 != Asignado, grouped by mat2 + setmana)
        bare_agg = defaultdict(int)
        # populated later via se_rows; stored in dashboard call separately

        fab_rows = sorted(
            [{"se_ref": k[0], "semana": k[1], "qty": v} for k, v in fab_agg.items() if v > 0],
            key=lambda x: (x["semana"], x["se_ref"])
        )
        result[fam] = {"pct": pct, "fab_rows": fab_rows}
    return result

def compute_dashboard_bare(se_rows, inventory):
    """
    Returns {family: [{se_ref, semana, qty}]} for bare stent deficit.
    ANGIOLIT: SRSI (SE32) needs SE21 bare, offset=3 weeks.
    ICOVER:   SE23 (encapsulado) needs SE221 bare, offset=1 week.
    Uses running pool: stock + bare planning lots with semana <= cutoff.
    """
    import copy

    # Build bare lots (SE21, SE221) sorted by semana — mutable for pool consumption
    bare_lots = defaultdict(list)
    for r in se_rows:
        art = r.get("article", "")
        sem = r.get("setmana")
        qty = r.get("peces_lot") or 0
        if (art and sem and qty > 0 and not art.endswith("_T")
                and (art.startswith("SE21") or art.startswith("SE221"))):
            bare_lots[art].append([sem, qty])
    for art in bare_lots:
        bare_lots[art].sort(key=lambda x: x[0])

    # Running stock of bare stents (SE21 and SE221 only, no _T)
    running = defaultdict(int)
    for art, qty in inventory.items():
        if not art.endswith("_T") and (art.startswith("SE21") or art.startswith("SE221")):
            running[art] = qty

    def _assign(demands, offset, deficit_agg):
        """Process demands sorted by semana, consuming running pool."""
        for sem, mat2_art, qty in sorted(demands, key=lambda x: x[0] or 9999):
            cutoff = sem - offset
            remaining = qty
            # From stock
            take = min(running[mat2_art], remaining)
            running[mat2_art] -= take
            remaining -= take
            # From bare planning lots with semana <= cutoff
            for lot in bare_lots.get(mat2_art, []):
                if lot[0] > cutoff:
                    break
                take = min(lot[1], remaining)
                lot[1] -= take
                remaining -= take
                if remaining == 0:
                    break
            if remaining > 0:
                deficit_agg[(mat2_art, sem)] += remaining

    # ANGIOLIT: SRSI lots → SE21, offset 3
    srsi = [(r["setmana"], r["material2"], r["peces_lot"] or 0)
            for r in se_rows
            if r.get("familia") == "SRSI"
            and r.get("disponibilitat_mat2") != STATUS_ASIGNADO
            and r.get("material2") and r.get("setmana") and (r.get("peces_lot") or 0) > 0]
    ang_deficit = defaultdict(int)
    _assign(srsi, 3, ang_deficit)

    # ICOVER: SE23 lots → SE221, offset 1
    enc = [(r["setmana"], r["material2"], r["peces_lot"] or 0)
           for r in se_rows
           if r.get("article", "").startswith("SE23")
           and r.get("disponibilitat_mat2") != STATUS_ASIGNADO
           and (r.get("material2") or "").startswith("SE221")
           and r.get("setmana") and (r.get("peces_lot") or 0) > 0]
    ico_deficit = defaultdict(int)
    _assign(enc, 1, ico_deficit)

    out = {}
    if ang_deficit:
        out["ANGIOLIT"] = sorted(
            [{"se_ref": k[0], "semana": k[1], "qty": v} for k, v in ang_deficit.items() if v > 0],
            key=lambda x: (x["semana"], x["se_ref"])
        )
    if ico_deficit:
        out["ICOVER"] = sorted(
            [{"se_ref": k[0], "semana": k[1], "qty": v} for k, v in ico_deficit.items() if v > 0],
            key=lambda x: (x["semana"], x["se_ref"])
        )
    return out

# ── CADUCIDADES DES ──────────────────────────────────────────────────────────
def compute_caducidades(inv_rows, pa_rows, weeks_20):
    """
    20-week matrix of SE32 lots expiring without enough non-Asignado PA demand.
    Returns {article: {semana: uncovered_qty}}
    """
    # Collect expiring SE32 lots grouped by (article, expiry_week)
    expiry = defaultdict(lambda: defaultdict(int))  # {art: {week: qty}}
    for r in inv_rows:
        art = str(r.get("articulo", ""))
        if not art.startswith("SE32"):
            continue
        cad = r.get("dia_cad")
        w   = date_to_week(cad)
        if w is None or w not in weeks_20:
            continue
        sta = r.get("cod_sta", 0)
        qty = (r.get("qnt_libre") or 0) if sta == 0 else (r.get("qnt_fis") or 0)
        if qty > 0:
            expiry[art][w] += qty

    # Non-Asignado PA demand per SE32 article per semana
    pa_dem = defaultdict(lambda: defaultdict(int))
    for r in pa_rows:
        if r.get("disponibilitat_mat2") != STATUS_ASIGNADO:
            se = r.get("material2", "")
            s  = r.get("setmana")
            q  = r.get("peces_lot") or 0
            if se and se.startswith("SE32") and s:
                pa_dem[se][s] += q

    result = {}
    for art, weeks_exp in expiry.items():
        art_result = {}
        # Sort expiry weeks; attribute demand greedily from earliest expiry
        remaining_demand = defaultdict(int)
        for s, q in pa_dem[art].items():
            remaining_demand[s] = q

        for w_exp in sorted(weeks_exp):
            qty_exp = weeks_exp[w_exp]
            # Available PA demand with semana < w_exp
            usable = sum(q for s, q in remaining_demand.items() if s < w_exp)
            uncovered = max(0, qty_exp - usable)
            if uncovered > 0:
                art_result[w_exp] = uncovered
                # Consume demand (FIFO by week)
                to_consume = qty_exp - uncovered
                for s in sorted(remaining_demand):
                    if s >= w_exp or to_consume <= 0:
                        break
                    take = min(remaining_demand[s], to_consume)
                    remaining_demand[s] -= take
                    to_consume -= take
            else:
                # consume qty_exp from earliest demand
                to_consume = qty_exp
                for s in sorted(remaining_demand):
                    if s >= w_exp or to_consume <= 0:
                        break
                    take = min(remaining_demand[s], to_consume)
                    remaining_demand[s] -= take
                    to_consume -= take

        if art_result:
            result[art] = art_result
    return result

# ── CADUCIDADES LIST (all expiring SE32 lots, chronological) ─────────────────
def build_cad_list(inv_rows, weeks_20):
    """Return [{ref, week}] for all SE32 lots expiring in weeks_20, sorted by (week, ref)."""
    entries = []
    for r in inv_rows:
        art = str(r.get("articulo", ""))
        if not art.startswith("SE32"):
            continue
        cad = r.get("dia_cad")
        w   = date_to_week(cad)
        if w is None or w not in weeks_20:
            continue
        sta = r.get("cod_sta", 0)
        qty = (r.get("qnt_libre") or 0) if sta == 0 else (r.get("qnt_fis") or 0)
        if qty > 0:
            entries.append({"ref": art, "week": w, "qty": qty})
    entries.sort(key=lambda x: (x["week"], x["ref"]))
    return entries

# ── KPIs ─────────────────────────────────────────────────────────────────────
def compute_kpis(pa_enriched, weeks):
    """Per-family KPIs for the top summary."""
    kpis = []
    by_fam = defaultdict(list)
    for r in pa_enriched:
        fam = PA_TO_DASH.get(r.get("familia", ""))
        if fam:
            by_fam[fam].append(r)
    for fam in FAMILY_ORDER:
        rows = by_fam[fam]
        refs_total  = len(set(r.get("article") for r in rows))
        refs_riesgo = len(set(r.get("article") for r in rows if r["status"] == STATUS_FABRICAR))
        uds_fab     = sum(r.get("uds_fab", 0) for r in rows)
        kpis.append({
            "familia":            fam,
            "refs_total":         refs_total,
            "refs_riesgo":        refs_riesgo,
            "unidades_a_fabricar": uds_fab,
        })
    return kpis

# ── TOP 15 DEFICITS ──────────────────────────────────────────────────────────
def compute_top(matrix, weeks):
    """Top 15 biggest negative coverage values across all families."""
    entries = []
    for fam, layers in matrix.items():
        for title, rows in layers.items():
            for row in rows:
                for idx, cob_val in enumerate(row["cob"]):
                    if cob_val < 0:
                        entries.append({
                            "familia": fam,
                            "capa":    title,
                            "ref":     row["ref"],
                            "medida":  row["medida"],
                            "deficit": abs(cob_val),
                            "semana":  weeks[idx],
                        })
    entries.sort(key=lambda x: x["deficit"], reverse=True)
    return entries[:15]

# ── PLANNING SE ROWS ─────────────────────────────────────────────────────────
def build_se_tab_rows(se_rows):
    """Return selected columns for Planning SE tab."""
    out = []
    for r in se_rows:
        art = r.get("article")
        if not art:
            continue
        fam = se_to_family(art) or r.get("familia", "")
        if not fam:
            continue
        fec = r.get("fec_ent_ped")
        fec_str = str(fec)[:10] if fec else ""
        dat = r.get("data_inici")
        dat_str = str(dat)[:10] if dat else ""
        out.append({
            "article":    art,
            "medidas":    r.get("medidas", ""),
            "familia":    r.get("familia", ""),
            "lot":        r.get("lot", ""),
            "setmana":    r.get("setmana"),
            "fabrica":    r.get("fabrica", ""),
            "peces_lot":  r.get("peces_lot") or 0,
            "fase_actual": r.get("fase_actual", ""),
            "data_inici": dat_str,
            "txt":        r.get("txt", "") or "",
        })
    return out

# ── INVENTARI TAB ─────────────────────────────────────────────────────────────
def build_inv_tab(inv_rows, fases_medida):
    """Return {factory: [{ref, medida, units}]} for inventory tab."""
    fases_medida = fases_medida or {}

    def is_obsolete(art):
        """SE21001-SE21021 are obsolete bare stents, exclude from inventory."""
        if art.startswith("SE21"):
            try:
                num = int("".join(filter(str.isdigit, art[4:])))
                return num <= 21
            except ValueError:
                pass
        return False

    # Stock qty from inventory (may be partial - some refs might have 0)
    stock = defaultdict(int)
    for r in inv_rows:
        art = str(r.get("articulo", ""))
        if is_obsolete(art) or art.endswith("_T"):
            continue
        sta = r.get("cod_sta", 0)
        qty = (r.get("qnt_libre") or 0) if sta == 0 else (r.get("qnt_fis") or 0)
        if qty > 0:
            stock[art] += qty

    # Source of refs: Fases_SE col G (same column as medidas) — excludes _T by design
    result = {}
    for fac in FACTORY_ORDER:
        arts = sorted(
            art for art in fases_medida.keys()
            if not art.endswith("_T") and not is_obsolete(art) and art_to_factory(art) == fac
        )
        rows = [{"ref": art, "medida": get_fases_medida(art, fases_medida), "units": stock.get(art, 0)}
                for art in arts]
        result[fac] = rows
    return result

# ── HTML GENERATION ──────────────────────────────────────────────────────────
def fmt_num(v):
    """Format number with thousands separator."""
    if v is None:
        return ""
    return f"{int(v):,}".replace(",", ".")

def generate_html(pa_enriched, se_rows, inv_rows, weeks, week_now, fases_articles=None, fases_medida=None):
    fases_medida = fases_medida or {}
    inventory  = build_inventory(inv_rows)
    matrix     = compute_family_matrix(pa_enriched, se_rows, inventory, weeks, fases_articles, fases_medida)
    dash       = compute_dashboard(pa_enriched)
    dash_bare  = compute_dashboard_bare(se_rows, inventory)
    kpis       = compute_kpis(pa_enriched, weeks)
    top        = compute_top(matrix, weeks)
    weeks_20   = list(range(week_now, week_now + 20))
    caducidades= compute_caducidades(inv_rows, pa_enriched, weeks_20)
    cad_list   = build_cad_list(inv_rows, weeks_20)
    se_tab     = build_se_tab_rows(se_rows)
    inv_tab    = build_inv_tab(inv_rows, fases_medida)

    # ── CSS ──────────────────────────────────────────────────────────────────
    css = """
* { box-sizing: border-box; margin: 0; padding: 0; }
html,body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif;
            background: #F4F6FA; color: #1F2937; height: 100%; }
body { display: flex; flex-direction: column; min-height: 100vh; }

header { background: linear-gradient(135deg, #2E3192 0%, #E91E8C 100%);
         color: white; padding: 16px 28px; box-shadow: 0 2px 8px rgba(0,0,0,.15); }
header h1 { font-size: 20px; font-weight: 600; letter-spacing: .3px; }
header .meta { font-size: 13px; opacity: .85; margin-top: 4px; }

nav.tabs { background: white; padding: 0 24px; border-bottom: 1px solid #E5E7EB;
           display: flex; gap: 2px; overflow-x: auto; flex-shrink: 0;
           position: sticky; top: 0; z-index: 10; }
nav.tabs button { background: none; border: none; padding: 14px 18px;
                  font-family: inherit; font-size: 13px; font-weight: 500;
                  color: #6B7280; cursor: pointer; border-bottom: 3px solid transparent;
                  transition: all .15s; white-space: nowrap; }
nav.tabs button:hover  { color: #2E3192; background: #F9FAFB; }
nav.tabs button.active { color: #2E3192; border-bottom-color: #E91E8C; font-weight: 600; }

main { padding: 24px 28px; flex: 1; }
.tab-panel { display: none; }
.tab-panel.active { display: block; }

.card { background: white; border-radius: 8px; padding: 20px; margin-bottom: 20px;
        box-shadow: 0 1px 3px rgba(0,0,0,.06); border: 1px solid #E5E7EB; }
.card h2 { font-size: 15px; font-weight: 600; margin-bottom: 14px; color: #2E3192;
           text-transform: uppercase; letter-spacing: .5px; }

/* Tables */
table { width: 100%; border-collapse: collapse; font-size: 12.5px; table-layout: fixed; }
thead th { background: #2E3192; color: white; padding: 10px 8px;
           text-align: center; font-weight: 600; font-size: 11.5px;
           position: sticky; top: 0; z-index: 2; overflow: hidden; }
tbody td { padding: 7px 8px; border-bottom: 1px solid #E5E7EB;
           vertical-align: middle; text-align: center; overflow: hidden; text-overflow: ellipsis; }
.planning-table { table-layout: auto; }
.planning-table td { overflow: visible; text-overflow: clip; white-space: nowrap; }
tbody tr:hover { background: #F9FAFB; }
.text-left { text-align: left !important; }
.font-bold { font-weight: 600; }
tr.section td { background: #D9E1F2; font-weight: 600; color: #1F3864;
                padding: 10px 12px; font-size: 12px; text-align: left; }
.scroll-x { overflow-x: auto; }
.scroll-y { overflow-y: auto; max-height: calc(100vh - 280px); }

/* Coverage cells */
td.cob { text-align: center; font-variant-numeric: tabular-nums; padding: 6px 4px; }
.cob-circle { display: inline-flex; min-width: 26px; height: 26px; border-radius: 13px;
              background: #FFC7CE; color: #9C0006; align-items: center;
              justify-content: center; font-weight: 700; font-size: 11px; padding: 0 5px; }

/* Deficit format (shared) */
.deficit-val { font-weight: 700; color: #9C0006; }

/* Status badges */
.badge { display: inline-block; padding: 3px 10px; border-radius: 12px;
         font-size: 11px; font-weight: 600; }
.badge-asignado  { background: #DDEBF7; color: #1F4E78; }
.badge-stock     { background: #C6EFCE; color: #006100; }
.badge-fab       { background: #FFEB9C; color: #9C6500; }
.badge-fabricar  { background: #FFC7CE; color: #9C0006; }

/* Filters */
.filters { display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 16px; align-items: center; }
.filters label { font-size: 12px; color: #4B5563; font-weight: 500; }
.filters select, .filters input { padding: 6px 10px; border: 1px solid #D1D5DB;
  border-radius: 6px; font-size: 13px; font-family: inherit; background: white; }
.filters input[type=text] { width: 180px; }

/* ── DASHBOARD ── */
.dash-grid { display: flex; gap: 16px; overflow-x: auto; align-items: flex-start; flex-wrap: wrap; }
.dash-col  { flex: 1; min-width: 170px; background: white; border-radius: 8px;
             border: 1px solid #E5E7EB; padding: 12px 10px;
             box-shadow: 0 1px 3px rgba(0,0,0,.06); }
.dash-fam-name { font-size: 12px; font-weight: 700; color: #2E3192; text-align: center;
                 text-transform: uppercase; letter-spacing: .5px; margin-bottom: 4px; }
.dash-table    { font-size: 11.5px; width: 100%; border-collapse: collapse; table-layout: auto; }
.dash-table th { background: #F3F4F6; color: #6B7280; font-size: 10.5px; font-weight: 600;
                 padding: 5px 4px; text-align: center; border-bottom: 1px solid #E5E7EB; }
.dash-table td { padding: 4px; text-align: center; border-bottom: 1px solid #F3F4F6;
                 font-variant-numeric: tabular-nums; }
.dash-table td.ref { text-align: left; font-family: monospace; font-size: 10.5px; }
.dash-sep td   { background: #EEF2FF; color: #3730A3; font-size: 10px; font-weight: 600;
                 padding: 3px 4px; text-align: center; }

/* ── INVENTARI BOXES ── */
.inv-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap: 10px; }
.inv-box  { background: white; border-radius: 8px; border: 1px solid #E5E7EB;
            box-shadow: 0 1px 3px rgba(0,0,0,.06); overflow: hidden; }
.inv-box-header { background: linear-gradient(135deg,#2E3192,#6B3FA0); color: white; padding: 6px 8px;
                  font-size: 11px; font-weight: 700; text-transform: uppercase; }
.inv-box table { font-size: 11px; table-layout: auto; width: 100%; }
.inv-box thead th { background: #EDE8F5; color: #2E3192; font-size: 10px; }
.inv-box tbody td { padding: 4px 6px; }

/* ── FAMILY BADGES ── */
.fam-badge        { display:inline-block; padding:2px 8px; border-radius:10px; font-size:11px; font-weight:600; }
.fam-angiolit     { background:#C9B1D9; color:#4A1B7A; }
.fam-angiolit-bare{ background:#E8D5F5; color:#6B3A99; }
.fam-icover       { background:#FFBCBC; color:#8B0000; }
.fam-icover-bare  { background:#FFD9D9; color:#CC0000; }
.fam-restorer     { background:#C6F6D5; color:#1A6B3F; }
.fam-naviscor     { background:#BFDBFE; color:#1E3A8A; }
.fam-ivolutio     { background:#FFEB9C; color:#9C6500; }
.fam-intercep     { background:#D1FAE5; color:#065F46; }

/* ── FAMILY TAB PANELS ── */
.fam-tab-header { padding: 18px 24px 14px; margin: -24px -28px 20px; }
.fam-tab-title  { font-size: 20px; font-weight: 700; color: white; letter-spacing: .3px; }

/* Per-family header + section title + table header colors */
.fam-tab-ivolutio .fam-tab-header { background: #FFEB9C; }
.fam-tab-ivolutio .fam-tab-title  { color: #000; }
.fam-tab-ivolutio thead th        { background: #FFEB9C; color: #000; }
.fam-tab-ivolutio .card h2        { background: #FFEB9C; color: #000; font-weight: 700; }
.fam-tab-ivolutio .card           { border-top: 3px solid #FFEB9C; }

.fam-tab-angiolit .fam-tab-header { background: #C9B1D9; }
.fam-tab-angiolit .fam-tab-title  { color: #000; }
.fam-tab-angiolit thead th        { background: #C9B1D9; color: #000; }
.fam-tab-angiolit .card h2        { background: #C9B1D9; color: #000; font-weight: 700; }
.fam-tab-angiolit .card           { border-top: 3px solid #C9B1D9; }

.fam-tab-naviscor .fam-tab-header { background: #BFDBFE; }
.fam-tab-naviscor .fam-tab-title  { color: #000; }
.fam-tab-naviscor thead th        { background: #BFDBFE; color: #000; }
.fam-tab-naviscor .card h2        { background: #BFDBFE; color: #000; font-weight: 700; }
.fam-tab-naviscor .card           { border-top: 3px solid #BFDBFE; }

.fam-tab-restorer .fam-tab-header { background: #1A6B3A; }
.fam-tab-restorer thead th        { background: #1A6B3A; color: #fff; }
.fam-tab-restorer .card h2        { background: #1A6B3A; color: #fff; font-weight: 700; }
.fam-tab-restorer .card           { border-top: 3px solid #1A6B3A; }

.fam-tab-icover .fam-tab-header   { background: #FFBCBC; }
.fam-tab-icover .fam-tab-title    { color: #000; }
.fam-tab-icover thead th          { background: #FFBCBC; color: #000; }
.fam-tab-icover .card h2          { background: #FFBCBC; color: #000; font-weight: 700; }
.fam-tab-icover .card             { border-top: 3px solid #FFBCBC; }

.fam-tab-intercep .fam-tab-header { background: #D1FAE5; }
.fam-tab-intercep .fam-tab-title  { color: #000; }
.fam-tab-intercep thead th        { background: #D1FAE5; color: #000; }
.fam-tab-intercep .card h2        { background: #D1FAE5; color: #000; font-weight: 700; }
.fam-tab-intercep .card           { border-top: 3px solid #D1FAE5; }

/* Stock bar column */
.stock-bar-cell  { padding: 4px 8px !important; min-width: 80px; width: 110px; }
.stock-bar-track { background: #F3F4F6; border-radius: 4px; height: 12px; width: 100%; }
.stock-bar-fill  { height: 12px; border-radius: 4px; min-width: 2px; }

/* Dashboard percentage bar */
.dash-pct-wrap  { background: #E5E7EB; border-radius: 8px; height: 26px; overflow: hidden; margin-bottom: 8px; }
.dash-pct-fill  { background: linear-gradient(135deg,#2E3192,#E91E8C); height: 100%;
                  border-radius: 8px; display: flex; align-items: center; min-width: 36px;
                  transition: width .4s; }
.dash-pct-label { color: white; font-size: 13px; font-weight: 700; padding: 0 8px; white-space: nowrap; }

/* ── CADUCIDADES ── */
.cad-scroll { overflow-x: auto; }
.cad-table  { font-size: 11.5px; border-collapse: collapse; table-layout: auto; width: auto; }
.cad-table th { background: #2E3192; color: white; padding: 8px 10px; text-align: center;
                font-size: 11px; white-space: nowrap; }
.cad-table td { padding: 6px 10px; border: 1px solid #E5E7EB; text-align: center;
                min-width: 52px; overflow: visible; text-overflow: clip; white-space: nowrap; }
.cad-table td.art { text-align: left; font-family: monospace; font-size: 11px;
                    background: #F9FAFB; font-weight: 600; white-space: nowrap;
                    min-width: 90px; }
"""

    # ── JS ────────────────────────────────────────────────────────────────────
    # Build DATA object
    data_obj = {
        "semana_actual": week_now,
        "weeks":         weeks,
        "kpis":          kpis,
        "top":           top,
        "dash":          dash,
        "dash_bare":     dash_bare,
        "matrix":        {fam: {title: rows for title, rows in layers.items()} for fam, layers in matrix.items()},
        "pa_rows":       [_pa_row_for_js(r) for r in pa_enriched],
        "se_tab":        se_tab,
        "inv_tab":        inv_tab,
        "factory_order":  FACTORY_ORDER,
        "factory_labels": FACTORY_LABELS,
        "caducidades":    {art: {str(w): v for w, v in wd.items()} for art, wd in caducidades.items()},
        "cad_list":       cad_list,
        "cad_weeks":      weeks_20,
        "fam_order":      FAMILY_ORDER,
    }
    data_json = json.dumps(data_obj, ensure_ascii=False, default=str)

    js = r"""
const FAM_NAMES = {
  ICOVER:'ICOVER (incl. TUVA)', ANGIOLIT:'ANGIOLITE (incl. ANGIOBTK)',
  NAVISCOR:'NAVISCORE', IVOLUTIO:'IVOLUTION', RESTORER:'RESTORER', INTERCEP:'INTERCEPT'
};

function fmt(v){ return v==null?'':v.toLocaleString('es-ES'); }
function defVal(v){ return v>0?`<span class="deficit-val">${fmt(v)}</span>`:''; }

function famBadgePA(fam){
  const map = {ANGIOLIT:'fam-angiolit',ICOVER:'fam-icover',RESTORER:'fam-restorer',
               NAVISCOR:'fam-naviscor',IVOLUTIO:'fam-ivolutio',INTERCEP:'fam-intercep'};
  const cls = map[fam]||'';
  return cls?`<span class="fam-badge ${cls}">${fam}</span>`:fam;
}
function famBadgeSE(fam, art){
  art = art||'';
  let cls;
  if(art.startsWith('SE21'))       cls='fam-angiolit-bare';
  else if(art.startsWith('SE221')) cls='fam-icover-bare';
  else if(art.startsWith('SE220')) cls='fam-restorer';
  else if(art.startsWith('SE32'))  cls='fam-angiolit';
  else if(art.startsWith('SE23'))  cls='fam-icover';
  else if(art.startsWith('SE24'))  cls='fam-ivolutio';
  else if(art.startsWith('SE26'))  cls='fam-naviscor';
  else if(art.startsWith('SE28'))  cls='fam-intercep';
  else cls='';
  return cls?`<span class="fam-badge ${cls}">${fam}</span>`:fam;
}

// ── TABS ──
function showTab(id){
  document.querySelectorAll('.tab-panel').forEach(p=>p.classList.remove('active'));
  document.querySelectorAll('nav.tabs button').forEach(b=>b.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  document.querySelector(`[data-tab="${id}"]`).classList.add('active');
}

// ── DASHBOARD ──
function buildDashboard(){
  const grid = document.getElementById('dash-grid');
  grid.innerHTML = '';
  DATA.fam_order.forEach(fam => {
    const d    = DATA.dash[fam] || {};
    const bare = DATA.dash_bare[fam] || [];
    const pct  = d.pct || 0;
    const rows = d.fab_rows || [];

    let tableRows = '';
    rows.forEach(r => {
      tableRows += `<tr><td class="ref">${r.se_ref}</td><td>S${r.semana}</td><td>${fmt(r.qty)}</td></tr>`;
    });
    if(rows.length > 0 && bare.length > 0){
      tableRows += `<tr class="dash-sep"><td colspan="3">▸ Bare stent</td></tr>`;
    }
    bare.forEach(r => {
      tableRows += `<tr><td class="ref">${r.se_ref}</td><td>S${r.semana}</td><td>${fmt(r.qty)}</td></tr>`;
    });
    if(!tableRows) tableRows = `<tr><td colspan="3" style="color:#6B7280;font-size:11px;padding:8px">Sin pendientes</td></tr>`;

    grid.innerHTML += `
      <div class="dash-col">
        <div class="dash-fam-name">${FAM_NAMES[fam]||fam}</div>
        <div class="dash-pct-wrap"><div class="dash-pct-fill" style="width:${pct}%"><span class="dash-pct-label">${pct}%</span></div></div>
        <table class="dash-table">
          <thead><tr><th>SE</th><th>W</th><th>Qty</th></tr></thead>
          <tbody>${tableRows}</tbody>
        </table>
      </div>`;
  });
}

// ── FAMILY MATRIX ──
const FAM_BAR_COLOR = {
  IVOLUTIO:'#FFEB9C', ANGIOLIT:'#C9B1D9', NAVISCOR:'#BFDBFE',
  RESTORER:'#1A6B3A', ICOVER:'#FFBCBC',   INTERCEP:'#D1FAE5'
};
function buildMatrix(fam){
  const container = document.getElementById('matrix-'+fam);
  if(!container) return;
  const barColor = FAM_BAR_COLOR[fam] || '#2E3192';
  const layers = DATA.matrix[fam];
  const weeks  = DATA.weeks;
  let html = '';
  Object.entries(layers).forEach(([title, rows]) => {
    const isBare = rows[0] && rows[0].is_bare;
    html += `<div class="card"><h2>${title}</h2><div class="scroll-x"><div class="scroll-y">`;
    const maxStock = Math.max(...rows.map(r => r.stock || 0), 1);
    html += `<table><thead><tr>
      <th style="text-align:left;width:90px">Ref</th>
      <th style="width:110px">Medida</th>
      <th>Stock</th><th style="width:110px"></th><th>En curso</th><th>Dem 6W</th>`;
    weeks.forEach(w => html += `<th>S${w}</th>`);
    html += `</tr></thead><tbody>`;
    rows.forEach(row => {
      const barPct = maxStock > 0 ? Math.round((row.stock||0)/maxStock*100) : 0;
      html += `<tr>
        <td class="text-left font-bold">${row.ref}</td>
        <td>${row.medida}</td>
        <td>${fmt(row.stock)}</td>
        <td class="stock-bar-cell"><div class="stock-bar-track"><div class="stock-bar-fill" style="width:${barPct}%;background:${barColor}"></div></div></td>
        <td>${fmt(row.en_curso)}</td>
        <td>${fmt(row.dem_6w)}</td>`;
      row.cob.forEach(v => {
        const txt = v < 0 ? `<span class="cob-circle">${fmt(Math.abs(v))}</span>` : '';
        html += `<td class="cob">${txt}</td>`;
      });
      html += `</tr>`;
    });
    html += `</tbody></table></div></div></div>`;
  });
  container.innerHTML = html;
}

// ── PLANNING PA ──
let paFilter = {familia:'', status:'', semana:'', search:''};
function buildPA(){
  const families = [...new Set(DATA.pa_rows.map(r=>r.familia_pa))].sort();
  const semanas  = [...new Set(DATA.pa_rows.map(r=>r.semana).filter(Boolean))].sort((a,b)=>a-b);
  const fSel = document.getElementById('pa-f-fam');
  const sSel = document.getElementById('pa-f-sem');
  const stSel= document.getElementById('pa-f-sta');
  if(fSel.options.length <= 1){
    families.forEach(f => { const o=document.createElement('option'); o.value=o.text=f; fSel.appendChild(o); });
    semanas.forEach(s  => { const o=document.createElement('option'); o.value=o.text=s; sSel.appendChild(o); });
  }
  refreshPA();
}
function refreshPA(){
  const f = paFilter;
  const rows = DATA.pa_rows.filter(r => {
    if(f.familia && r.familia_pa !== f.familia) return false;
    if(f.status  && r.status   !== f.status)   return false;
    if(f.semana  && String(r.semana) !== f.semana) return false;
    if(f.search){
      const s = f.search.toLowerCase();
      if(!(r.article||'').toLowerCase().includes(s) &&
         !(r.medidas||'').toLowerCase().includes(s) &&
         !(String(r.lot||'')).includes(s)) return false;
    }
    return true;
  });
  const tbody = document.getElementById('pa-tbody');
  tbody.innerHTML = rows.map(r => {
    const badge = r.status === 'Asignado'      ? 'badge-asignado'
                : r.status === 'En stock'       ? 'badge-stock'
                : r.status === 'En fabricación' ? 'badge-fab'
                :                                 'badge-fabricar';
    const udsFab = r.uds_fab > 0 ? `<span class="deficit-val">${fmt(r.uds_fab)}</span>` : '';
    return `<tr>
      <td class="text-left">${r.article}</td>
      <td>${r.medidas}</td>
      <td>${famBadgePA(r.familia_pa)}</td>
      <td>${r.lot}</td>
      <td>${r.semana||''}</td>
      <td>${r.stent||''}</td>
      <td>${fmt(r.peces_lot)}</td>
      <td><span class="badge ${badge}">${r.status}</span></td>
      <td>${udsFab}</td>
    </tr>`;
  }).join('');
  document.getElementById('pa-count').textContent = `${rows.length} registros`;
}

// ── PLANNING SE ──
let seFilter = {familia:'', semana:'', search:''};
function buildSE(){
  const families = [...new Set(DATA.se_tab.map(r=>r.familia))].sort();
  const semanas  = [...new Set(DATA.se_tab.map(r=>r.setmana).filter(Boolean))].sort((a,b)=>a-b);
  const fSel = document.getElementById('se-f-fam');
  const sSel = document.getElementById('se-f-sem');
  if(fSel.options.length <= 1){
    families.forEach(f => { const o=document.createElement('option'); o.value=o.text=f; fSel.appendChild(o); });
    semanas.forEach(s  => { const o=document.createElement('option'); o.value=o.text=s; sSel.appendChild(o); });
  }
  refreshSE();
}
function refreshSE(){
  const f = seFilter;
  const rows = DATA.se_tab.filter(r => {
    if(f.familia && r.familia !== f.familia) return false;
    if(f.semana  && String(r.setmana) !== f.semana) return false;
    if(f.search){
      const s = f.search.toLowerCase();
      if(!(r.article||'').toLowerCase().includes(s) &&
         !(r.medidas||'').toLowerCase().includes(s) &&
         !(String(r.lot||'')).includes(s)) return false;
    }
    return true;
  });
  const tbody = document.getElementById('se-tbody');
  tbody.innerHTML = rows.map(r => {
    const cortando = r.data_inici ? '<span class="badge badge-fab">Cortando</span>' : '';
    return `<tr>
      <td class="text-left">${r.article}</td>
      <td>${r.medidas}</td>
      <td>${famBadgeSE(r.familia, r.article)}</td>
      <td>${r.lot||''}</td>
      <td>${r.setmana||''}</td>
      <td>${r.fabrica||''}</td>
      <td>${fmt(r.peces_lot)}</td>
      <td class="text-left">${r.fase_actual||''}</td>
      <td>${cortando}</td>
      <td class="text-left">${r.txt||''}</td>
    </tr>`;
  }).join('');
  document.getElementById('se-count').textContent = `${rows.length} registros`;
}

// ── INVENTARI ──
function buildInv(){
  const grid = document.getElementById('inv-grid');
  if(grid.innerHTML.trim()) return;
  let html = '';
  DATA.factory_order.forEach(fac => {
    const label = DATA.factory_labels[fac] || fac;
    const rows  = DATA.inv_tab[fac] || [];
    let trows = rows.map(r => `<tr>
      <td class="text-left" style="font-family:monospace">${r.ref}</td>
      <td>${r.medida}</td>
      <td>${fmt(r.units)}</td>
    </tr>`).join('') || '<tr><td colspan="3" style="color:#6B7280">Sin stock</td></tr>';
    html += `<div class="inv-box">
      <div class="inv-box-header">${label}</div>
      <table>
        <thead><tr><th style="text-align:left">Referencia</th><th>Medida</th><th>Uds</th></tr></thead>
        <tbody>${trows}</tbody>
      </table>
    </div>`;
  });
  grid.innerHTML = html;
}

// ── CADUCIDADES ──
function buildCaducidades(){
  const container = document.getElementById('cad-container');
  if(container.innerHTML.trim()) return;
  const weeks = DATA.cad_weeks;
  const cad   = DATA.caducidades;
  const arts  = Object.keys(cad).sort();

  // Matrix (sin cobertura)
  let html = '';
  if(!arts.length){
    html += '<p style="color:#6B7280;padding:16px">No se detectan caducidades problemáticas en las próximas 20 semanas.</p>';
  } else {
    html += '<div class="cad-scroll"><table class="cad-table"><thead><tr><th>Referencia</th>';
    weeks.forEach(w => html += `<th>S${w}</th>`);
    html += '</tr></thead><tbody>';
    arts.forEach(art => {
      html += `<tr><td class="art">${art}</td>`;
      weeks.forEach(w => {
        const v = cad[art][String(w)];
        html += `<td>${v ? `<span class="deficit-val">${fmt(v)}</span>` : ''}</td>`;
      });
      html += '</tr>';
    });
    html += '</tbody></table></div>';
  }

  // Lista cronológica de caducidades
  const list = DATA.cad_list || [];
  html += `<h3 style="margin-top:24px;margin-bottom:10px;font-size:13px;font-weight:700;
    color:#1F3864;text-transform:uppercase;letter-spacing:.5px">
    Listado cronológico de caducidades SE32</h3>`;
  if(!list.length){
    html += '<p style="color:#6B7280;font-size:12px">Sin caducidades en las próximas 20 semanas.</p>';
  } else {
    html += '<div class="cad-scroll"><table class="cad-table"><thead><tr>'
          + '<th style="text-align:left">Referencia</th><th>Semana</th><th>Uds</th>'
          + '</tr></thead><tbody>';
    list.forEach(e => {
      html += `<tr><td style="font-family:monospace;text-align:left">${e.ref}</td>`
            + `<td>S${e.week}</td><td>${fmt(e.qty)}</td></tr>`;
    });
    html += '</tbody></table></div>';
  }

  container.innerHTML = html;
}

// ── INIT ──
document.addEventListener('DOMContentLoaded', () => {
  showTab('tab-dashboard');
  buildDashboard();
  DATA.fam_order.forEach(fam => buildMatrix(fam));
  buildPA();
  buildSE();
  buildInv();
  buildCaducidades();

  document.getElementById('pa-f-fam').addEventListener('change', e => { paFilter.familia = e.target.value; refreshPA(); });
  document.getElementById('pa-f-sem').addEventListener('change', e => { paFilter.semana  = e.target.value; refreshPA(); });
  document.getElementById('pa-f-sta').addEventListener('change', e => { paFilter.status  = e.target.value; refreshPA(); });
  document.getElementById('pa-search').addEventListener('input',  e => { paFilter.search  = e.target.value; refreshPA(); });

  document.getElementById('se-f-fam').addEventListener('change', e => { seFilter.familia = e.target.value; refreshSE(); });
  document.getElementById('se-f-sem').addEventListener('change', e => { seFilter.semana  = e.target.value; refreshSE(); });
  document.getElementById('se-search').addEventListener('input',  e => { seFilter.search  = e.target.value; refreshSE(); });
});
"""

    # ── HTML BODY ────────────────────────────────────────────────────────────
    # Nav tabs
    tab_defs = [("tab-dashboard", "Resumen")]
    for fam in FAMILY_ORDER:
        tab_defs.append((f"tab-{fam.lower()}", FAMILY_NAMES[fam]))
    tab_defs += [
        ("tab-planning-pa", "Planning PA"),
        ("tab-planning-se", "Planning SE"),
        ("tab-inventari",   "Inventario"),
        ("tab-caducidades", "Caducidades DES"),
    ]
    nav_html = "".join(
        f'<button data-tab="{tid}" onclick="showTab(\'{tid}\')">{label}</button>'
        for tid, label in tab_defs
    )

    # KPI cards for top summary
    kpi_html = "".join(f"""
      <div class="kpi" style="background:white;border-radius:8px;padding:16px;border:1px solid #E5E7EB">
        <div style="font-size:11px;color:#6B7280;font-weight:600;text-transform:uppercase;letter-spacing:.5px">{k['familia']}</div>
        <div style="font-size:22px;font-weight:700;color:{'#9C0006' if k['refs_riesgo']>0 else '#006100'}">{k['refs_riesgo']}<span style="font-size:13px;color:#6B7280">/{k['refs_total']} en riesgo</span></div>
        <div style="font-size:12px;color:#6B7280;margin-top:4px">Fab: <b>{fmt_num(k['unidades_a_fabricar'])} uds</b></div>
      </div>""" for k in kpis)

    # Top-15 table
    top_html = "".join(f"""<tr>
      <td class="text-left">{t['familia']}</td>
      <td class="text-left">{t['ref']}</td>
      <td class="text-left">{t['medida']}</td>
      <td><span class="deficit-val">{fmt_num(t['deficit'])}</span></td>
      <td>{t['semana']}</td>
    </tr>""" for t in top)

    # PA tab columns header
    pa_thead = """<tr>
      <th style="text-align:left">Artículo</th><th>Medida</th><th>Familia</th>
      <th>Lot</th><th>S</th><th>Stent (SE)</th><th>Lot Qty</th>
      <th>Estado</th><th>Uds fabricar</th>
    </tr>"""

    # SE tab header
    se_thead = """<tr>
      <th style="text-align:left">Artículo</th><th>Medida</th><th>Familia</th>
      <th>Lot</th><th>S</th><th>Fábrica</th><th>Peces lot</th>
      <th style="text-align:left">Fase actual</th><th>Cortando</th><th style="text-align:left">Txt</th>
    </tr>"""

    # Status filter options for PA
    status_opts = "".join(f'<option value="{s}">{s}</option>'
                          for s in [STATUS_ASIGNADO, STATUS_EN_STOCK, STATUS_EN_FAB, STATUS_FABRICAR])

    # Family matrix panels
    matrix_panels = "".join(
        f'<div class="tab-panel fam-tab-{fam.lower()}" id="tab-{fam.lower()}">'
        f'<div class="fam-tab-header"><span class="fam-tab-title">{FAMILY_NAMES[fam]}</span></div>'
        f'<div id="matrix-{fam}"></div>'
        f'</div>'
        for fam in FAMILY_ORDER
    )

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<title>Indicadores Stock Stents · Semana {week_now}</title>
<style>{css}</style>
</head>
<body>
<header>
  <h1>Indicadores Stock Stents</h1>
  <div class="meta">Semana actual: <b>{week_now}</b> · Ventana: semanas {weeks[0]}–{weeks[-1]} · Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}</div>
</header>
<nav class="tabs">{nav_html}</nav>
<main>

<!-- DASHBOARD -->
<div class="tab-panel active" id="tab-dashboard">
  <div class="card" style="margin-bottom:16px">
    <h2>KPIs por familia</h2>
    <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px">{kpi_html}</div>
  </div>
  <div class="card">
    <h2>Estado de fabricación por familia</h2>
    <div id="dash-grid" class="dash-grid"></div>
  </div>
</div>

<!-- FAMILY TABS -->
{matrix_panels}

<!-- PLANNING PA -->
<div class="tab-panel" id="tab-planning-pa">
  <div class="card">
    <div class="filters">
      <label>Familia <select id="pa-f-fam"><option value="">Todas</option></select></label>
      <label>Semana <select id="pa-f-sem"><option value="">Todas</option></select></label>
      <label>Estado <select id="pa-f-sta"><option value="">Todos</option>{status_opts}</select></label>
      <label>Búsqueda <input id="pa-search" type="text" placeholder="Artículo / medida / lot..."></label>
      <span id="pa-count" style="font-size:12px;color:#6B7280;margin-left:auto"></span>
    </div>
    <div class="scroll-x"><div class="scroll-y">
      <table class="planning-table"><thead>{pa_thead}</thead><tbody id="pa-tbody"></tbody></table>
    </div></div>
  </div>
</div>

<!-- PLANNING SE -->
<div class="tab-panel" id="tab-planning-se">
  <div class="card">
    <div class="filters">
      <label>Familia <select id="se-f-fam"><option value="">Todas</option></select></label>
      <label>Semana <select id="se-f-sem"><option value="">Todas</option></select></label>
      <label>Búsqueda <input id="se-search" type="text" placeholder="Artículo / medida / lot..."></label>
      <span id="se-count" style="font-size:12px;color:#6B7280;margin-left:auto"></span>
    </div>
    <div class="scroll-x"><div class="scroll-y">
      <table class="planning-table"><thead>{se_thead}</thead><tbody id="se-tbody"></tbody></table>
    </div></div>
  </div>

</div>

<!-- INVENTARI -->
<div class="tab-panel" id="tab-inventari">
  <div class="card">
    <h2>Stock disponible por fábrica</h2>
    <div id="inv-grid" class="inv-grid"></div>
  </div>
</div>

<!-- CADUCIDADES DES -->
<div class="tab-panel" id="tab-caducidades">
  <div class="card">
    <h2>Caducidades DES sin cobertura · próximas 20 semanas</h2>
    <p style="font-size:12px;color:#6B7280;margin-bottom:12px">
      Solo se muestran referencias con unidades caducando sin órdenes PA no-asignadas suficientes.
    </p>
    <div id="cad-container"></div>
  </div>
</div>

</main>

<a href="datos.xlsx" download="indicadores_stents_S{week_now}.xlsx"
  title="Descargar datos en Excel"
  style="position:fixed;bottom:18px;right:18px;z-index:1000;display:block;padding:4px;opacity:.85;transition:opacity .2s;text-decoration:none"
  onmouseover="this.style.opacity=1" onmouseout="this.style.opacity=.85">
  <svg width="38" height="38" viewBox="0 0 48 48" xmlns="http://www.w3.org/2000/svg">
    <rect width="48" height="48" rx="6" fill="#1D6F42"/>
    <rect x="26" y="6" width="16" height="36" rx="2" fill="#185C37"/>
    <rect x="6" y="6" width="22" height="36" rx="2" fill="#21A366"/>
    <path d="M6 6h22v6H6z" fill="#107C41" opacity=".5"/>
    <text x="14" y="30" font-family="Arial,sans-serif" font-size="18" font-weight="900" fill="white">X</text>
    <line x1="27" y1="14" x2="42" y2="14" stroke="white" stroke-width="1" opacity=".4"/>
    <line x1="27" y1="20" x2="42" y2="20" stroke="white" stroke-width="1" opacity=".4"/>
    <line x1="27" y1="26" x2="42" y2="26" stroke="white" stroke-width="1" opacity=".4"/>
    <line x1="27" y1="32" x2="42" y2="32" stroke="white" stroke-width="1" opacity=".4"/>
  </svg>
</a>

<script>
const DATA = {data_json};
</script>
<script>
{js}
</script>
</body>
</html>"""
    return html

def _pa_row_for_js(r):
    return {
        "article":    r.get("article", ""),
        "medidas":    r.get("medidas", ""),
        "familia_pa": PA_TO_DASH.get(r.get("familia", ""), r.get("familia", "")),
        "lot":        r.get("lot", ""),
        "semana":     r.get("setmana"),
        "stent":      r.get("material2", ""),
        "peces_lot":  r.get("peces_lot") or 0,
        "status":     r.get("status", ""),
        "uds_fab":    r.get("uds_fab", 0),
    }

# ── EXCEL EXPORT ─────────────────────────────────────────────────────────────
def _xl_header(ws, cols, row=1):
    """Header row: dark blue background, white bold text."""
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor="2E3192")
    font = Font(color="FFFFFF", bold=True, size=10)
    for ci, label in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci, value=label)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

def _xl_section(ws, label, ncols, row):
    """Section separator row: light gray background."""
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor="E5E7EB")
    font = Font(bold=True, size=10, color="374151")
    c = ws.cell(row=row, column=1, value=label)
    c.fill = fill
    c.font = font
    c.alignment = Alignment(horizontal="left")
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    return row + 1

def _xl_autofit(ws):
    """Auto-fit column widths (capped at 40)."""
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=4)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

def generate_excel(pa_enriched, se_rows, inv_rows, weeks, week_now,
                   fases_articles=None, fases_medida=None):
    """Generate datos.xlsx with one sheet per HTML tab, clean data only."""
    fases_medida = fases_medida or {}
    inventory    = build_inventory(inv_rows)
    matrix       = compute_family_matrix(pa_enriched, se_rows, inventory, weeks,
                                         fases_articles, fases_medida)
    kpis         = compute_kpis(pa_enriched, weeks)
    top          = compute_top(matrix, weeks)
    weeks_20     = list(range(week_now, week_now + 20))
    cad_list     = build_cad_list(inv_rows, weeks_20)
    se_tab       = build_se_tab_rows(se_rows)
    inv_tab      = build_inv_tab(inv_rows, fases_medida)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    # ── Resumen ───────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Resumen")
    kpi_cols = ["Familia", "Refs Total", "Refs Riesgo", "Uds a Fabricar"]
    _xl_header(ws, kpi_cols, row=1)
    for i, k in enumerate(kpis, 2):
        ws.cell(i, 1, FAMILY_NAMES.get(k["familia"], k["familia"]))
        ws.cell(i, 2, k["refs_total"])
        ws.cell(i, 3, k["refs_riesgo"])
        ws.cell(i, 4, k["unidades_a_fabricar"])

    # Top 15 déficits (con separador)
    sep_row = len(kpis) + 3
    top_cols = ["Familia", "Capa", "Referencia", "Medida", "Déficit", "Semana"]
    _xl_section(ws, "Top 15 Déficits de cobertura", len(top_cols), sep_row)
    _xl_header(ws, top_cols, row=sep_row + 1)
    for i, t in enumerate(top, sep_row + 2):
        ws.cell(i, 1, FAMILY_NAMES.get(t["familia"], t["familia"]))
        ws.cell(i, 2, t["capa"])
        ws.cell(i, 3, t["ref"])
        ws.cell(i, 4, t["medida"])
        ws.cell(i, 5, t["deficit"])
        ws.cell(i, 6, f"S{t['semana']}")
    _xl_autofit(ws)

    # ── Pestañas de familia (6 hojas) ─────────────────────────────────────────
    week_headers = [f"S{w}" for w in weeks]
    matrix_cols  = ["Referencia", "Medida", "Stock", "En Curso", "Dem 6W"] + week_headers

    for fam in FAMILY_ORDER:
        ws = wb.create_sheet(FAMILY_NAMES.get(fam, fam))
        _xl_header(ws, matrix_cols, row=1)
        row = 2
        for title, rows in matrix.get(fam, {}).items():
            row = _xl_section(ws, title, len(matrix_cols), row)
            for r in rows:
                ws.cell(row, 1, r["ref"])
                ws.cell(row, 2, r["medida"])
                ws.cell(row, 3, r["stock"])
                ws.cell(row, 4, r["en_curso"])
                ws.cell(row, 5, r["dem_6w"])
                for ci, cob_val in enumerate(r["cob"], 6):
                    ws.cell(row, ci, cob_val)
                row += 1
        _xl_autofit(ws)

    # ── Planning PA ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Planning PA")
    pa_cols = ["Artículo", "Medidas", "Familia", "Lot", "Semana",
               "Stent SE", "Peces/Lot", "Status", "Uds Fab"]
    _xl_header(ws, pa_cols, row=1)
    for i, r in enumerate(pa_enriched, 2):
        ws.cell(i, 1, r.get("article", ""))
        ws.cell(i, 2, r.get("medidas", ""))
        ws.cell(i, 3, r.get("familia", ""))
        ws.cell(i, 4, r.get("lot", ""))
        ws.cell(i, 5, r.get("setmana"))
        ws.cell(i, 6, r.get("material2", ""))
        ws.cell(i, 7, r.get("peces_lot") or 0)
        ws.cell(i, 8, r.get("status", ""))
        ws.cell(i, 9, r.get("uds_fab", 0))
    _xl_autofit(ws)

    # ── Planning SE ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Planning SE")
    se_cols = ["Artículo", "Medidas", "Familia", "Lot", "Semana",
               "Fábrica", "Peces/Lot", "Fase Actual", "Data Inici", "Comentario"]
    _xl_header(ws, se_cols, row=1)
    for i, r in enumerate(se_tab, 2):
        ws.cell(i, 1, r.get("article", ""))
        ws.cell(i, 2, r.get("medidas", ""))
        ws.cell(i, 3, r.get("familia", ""))
        ws.cell(i, 4, r.get("lot", ""))
        ws.cell(i, 5, r.get("setmana"))
        ws.cell(i, 6, r.get("fabrica", ""))
        ws.cell(i, 7, r.get("peces_lot") or 0)
        ws.cell(i, 8, r.get("fase_actual", ""))
        ws.cell(i, 9, r.get("data_inici", ""))
        ws.cell(i, 10, r.get("txt", ""))
    _xl_autofit(ws)

    # ── Inventario ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Inventario")
    inv_cols = ["Fábrica", "Referencia", "Medida", "Stock"]
    _xl_header(ws, inv_cols, row=1)
    row = 2
    for fac in FACTORY_ORDER:
        rows = inv_tab.get(fac, [])
        if not rows:
            continue
        row = _xl_section(ws, FACTORY_LABELS.get(fac, fac), len(inv_cols), row)
        for r in rows:
            ws.cell(row, 1, fac)
            ws.cell(row, 2, r["ref"])
            ws.cell(row, 3, r["medida"])
            ws.cell(row, 4, r["units"])
            row += 1
    _xl_autofit(ws)

    # ── Caducidades DES ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Caducidades DES")
    cad_cols = ["Referencia", "Semana", "Qty"]
    _xl_header(ws, cad_cols, row=1)
    for i, r in enumerate(cad_list, 2):
        ws.cell(i, 1, r["ref"])
        ws.cell(i, 2, f"S{r['week']}")
        ws.cell(i, 3, r["qty"])
    _xl_autofit(ws)

    wb.save(EXCEL_OUT)
    print(f"Excel generado: {EXCEL_OUT} ({EXCEL_OUT.stat().st_size // 1024} KB)")

# ── GIT PUSH ─────────────────────────────────────────────────────────────────
def git_push(week_now):
    cmds = [
        ["git", "-C", str(REPO_PATH), "add", "index.html", "datos.xlsx"],
        ["git", "-C", str(REPO_PATH), "commit", "-m", f"Auto-update semana {week_now}"],
        ["git", "-C", str(REPO_PATH), "push"],
    ]
    for cmd in cmds:
        result = subprocess.run(cmd, capture_output=True, text=True)
        if result.returncode != 0 and "nothing to commit" not in result.stdout:
            print(f"Git warning: {result.stderr.strip()}")

# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("Cargando Excel...")
    pa_rows, se_rows, inv_rows, fases_rows, fases_medida = load_excel()

    week_now  = current_week()
    weeks     = list(range(week_now, week_now + 6))
    print(f"Semana actual: {week_now} | Ventana: {weeks}")

    # Todos los artículos SE existentes según Fases_SE (excluir _T)
    fases_articles = list(set(
        r.get("SE") for r in fases_rows
        if r.get("SE") and not str(r.get("SE")).endswith("_T")
    ))
    print(f"Artículos SE en Fases_SE: {len(fases_articles)} (sin _T)")

    print("Calculando inventario...")
    inventory = build_inventory(inv_rows)

    print("Calculando planning SE...")
    se_planning = build_se_planning(se_rows)

    print("Asignando lotes PA...")
    pa_enriched = run_pa_assignment(pa_rows, inventory, se_planning)

    print("Generando HTML...")
    html = generate_html(pa_enriched, se_rows, inv_rows, weeks, week_now, fases_articles, fases_medida)
    HTML_OUT.write_text(html, encoding="utf-8")
    print(f"HTML generado: {HTML_OUT} ({len(html)//1024} KB)")

    print("Generando Excel de datos...")
    generate_excel(pa_enriched, se_rows, inv_rows, weeks, week_now, fases_articles, fases_medida)

    print("Subiendo a GitHub...")
    git_push(week_now)
    print("Listo. https://mcasanovas23.github.io/Indicadors-planning/")

if __name__ == "__main__":
    main()
